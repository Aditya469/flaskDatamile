import os
import shutil
import csv
import socket
import pandas as pd
import logging
from reportlab.lib import colors
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import landscape, letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Spacer, KeepTogether, PageBreak
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import TableStyleInfo
from io import BytesIO
from io import StringIO
from itertools import groupby
from operator import attrgetter
from flask import current_app
from flask import Blueprint, render_template, request, flash, redirect, url_for, jsonify, send_file, make_response
from flask_login import login_required, current_user
from .models import SalesOrder, Picklist, StockPrice, CancelledList, WeeklyForecasts, WeeklyTotals, MonthlyForecasts, MonthlyTotals, GrandTotal
from .deleteSalesOrder import delete_sales_orders_and_picklist
from . import db
from datetime import datetime
from sqlalchemy import text, func, case, and_
from sqlalchemy.types import Integer
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.orm import Session
from contextlib import contextmanager
from project import create_app
from PyPDF2 import PdfMerger

main = Blueprint('main', __name__)

# Set the upload folder path
UPLOAD_FOLDER = "C:/Users/Rama/flaskDatamile/project/hse_storage"

extended_weeks_value = '0'

def get_server_info():
    hostname = socket.gethostname()
    ip_address = socket.gethostbyname(hostname)
    port = 90  # Default Flask port, change if you're using a different port
    return f"{ip_address}:{port}"

@main.context_processor
def inject_server_info():
    return dict(server_info=get_server_info())

@main.route('/')
def index():
    selected_table = request.args.get('table', 'picklist')
    selected_account_id = request.args.get('account_id', 'All')
    selected_week = request.args.get('week', 'All')
    extended_weeks = request.args.getlist('extended_weeks')
    show_further_weeks = request.args.get('show_further_weeks', 'false').lower() == 'true'

    # Define default weeks
    default_weeks = ['Arrears', 'Week - 1', 'Week - 2', 'Week - 3', 'Week - 4']

    # Define week_order for Picklist
    picklist_week_order = case(
        (Picklist.week == 'Arrears', 0),
        (Picklist.week == 'Week - 1', 1),
        (Picklist.week == 'Week - 2', 2),
        (Picklist.week == 'Week - 3', 3),
        (Picklist.week == 'Week - 4', 4),
        (Picklist.week == 'Week - 5', 5),
        (Picklist.week == 'Week - 6', 6),
        (Picklist.week == 'Week - 7', 7),
        (Picklist.week == 'Week - 8', 8),
        else_=9
    )

    # Define week_order for CancelledList
    cancelled_week_order = case(
        (CancelledList.week == 'Arrears', 0),
        (CancelledList.week == 'Week - 1', 1),
        (CancelledList.week == 'Week - 2', 2),
        (CancelledList.week == 'Week - 3', 3),
        (CancelledList.week == 'Week - 4', 4),
        (CancelledList.week == 'Week - 5', 5),
        (CancelledList.week == 'Week - 6', 6),
        (CancelledList.week == 'Week - 7', 7),
        (CancelledList.week == 'Week - 8', 8),
        else_=9
    )

    # Fetch Picklist items
    picklist_query = Picklist.query.order_by(
        Picklist.account_id.asc(),
        picklist_week_order,
        func.date_format(Picklist.required_date, '%d-%m-%Y').asc(),
        Picklist.required_day.asc(),
        Picklist.stock_code.asc()
    )

    # Fetch CancelledList items
    cancelled_query = CancelledList.query.order_by(
        CancelledList.account_id.asc(),
        cancelled_week_order,
        func.date_format(CancelledList.required_date, '%d-%m-%Y').asc(),
        CancelledList.required_day.asc(),
        CancelledList.stock_code.asc()
    )

    if selected_account_id != 'All':
        picklist_query = picklist_query.filter(Picklist.account_id == selected_account_id)
        cancelled_query = cancelled_query.filter(CancelledList.account_id == selected_account_id)

    if selected_week != 'All':
        picklist_query = picklist_query.filter(Picklist.week == selected_week)
        cancelled_query = cancelled_query.filter(CancelledList.week == selected_week)
    else:
        allowed_weeks = default_weeks + (extended_weeks if show_further_weeks else [])
        picklist_query = picklist_query.filter(Picklist.week.in_(allowed_weeks))
        cancelled_query = cancelled_query.filter(CancelledList.week.in_(allowed_weeks))

    picklist_items = picklist_query.all()
    cancelled_items = cancelled_query.all()

    account_ids = db.session.query(Picklist.account_id).union(
        db.session.query(CancelledList.account_id)
    ).distinct().order_by(Picklist.account_id).all()
    account_ids = [account_id[0] for account_id in account_ids]

    weeks = db.session.query(Picklist.week).union(
        db.session.query(CancelledList.week)
    ).distinct().order_by(Picklist.week).all()
    weeks = [week[0] for week in weeks]

    # Modify the required_dates_query
    picklist_dates = db.session.query(func.date_format(Picklist.required_date, '%d-%m-%Y').label('formatted_date'))
    cancelled_dates = db.session.query(func.date_format(CancelledList.required_date, '%d-%m-%Y').label('formatted_date'))
    required_dates_query = picklist_dates.union(cancelled_dates).distinct()

    if selected_account_id != 'All':
        picklist_dates = picklist_dates.filter(Picklist.account_id == selected_account_id)
        cancelled_dates = cancelled_dates.filter(CancelledList.account_id == selected_account_id)
        required_dates_query = picklist_dates.union(cancelled_dates).distinct()

    if selected_week != 'All':
        picklist_dates = picklist_dates.filter(Picklist.week == selected_week)
        cancelled_dates = cancelled_dates.filter(CancelledList.week == selected_week)
        required_dates_query = picklist_dates.union(cancelled_dates).distinct()
    else:
        # Filter by default weeks and extended weeks
        allowed_weeks = default_weeks + extended_weeks
        picklist_dates = picklist_dates.filter(Picklist.week.in_(allowed_weeks))
        cancelled_dates = cancelled_dates.filter(CancelledList.week.in_(allowed_weeks))
        required_dates_query = picklist_dates.union(cancelled_dates).distinct()

    # Fetch all dates
    required_dates = required_dates_query.all()
    required_dates = [date[0] for date in required_dates]

    # Custom sorting function
    def date_sort_key(date_string):
        return datetime.strptime(date_string, '%d-%m-%Y')

    # Sort the dates
    required_dates.sort(key=date_sort_key)

    return render_template('index.html',
                           picklist_items=picklist_items,
                           cancelled_items=cancelled_items,
                           account_ids=account_ids,
                           weeks=weeks,
                           required_dates=required_dates,
                           selected_table=selected_table,
                           selected_account_id=selected_account_id,
                           selected_week=selected_week,
                           extended_weeks=extended_weeks,
                           show_further_weeks=show_further_weeks)


class NumberedCanvas(canvas.Canvas):
    def __init__(self, *args, **kwargs):
        canvas.Canvas.__init__(self, *args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_number(num_pages)
            canvas.Canvas.showPage(self)
        canvas.Canvas.save(self)

    def draw_page_number(self, page_count):
        self.setFont("Helvetica", 9)
        self.drawRightString(
            self._pagesize[0] - 0.5*inch,
            0.5*inch,
            f"Page {self._pageNumber} of {page_count}"
        )

def header(canvas, doc, account_id):
    canvas.saveState()
    styles = getSampleStyleSheet()
    
    # Logo
    logo_path = os.path.join(current_app.root_path, 'static', 'logo', 'ecam_logo.jpg')
    if os.path.exists(logo_path):
        canvas.drawImage(logo_path, doc.leftMargin, doc.height + doc.topMargin - 0.75*inch, width=1*inch, height=0.75*inch)
    
    # Header text
    header_text = f"ECAM ENGINEERING LIMITED - PICKLIST - {account_id}"
    canvas.setFont("Helvetica-Bold", 14)
    canvas.drawCentredString(doc.width/2.0 + doc.leftMargin, doc.height + doc.topMargin - 0.3*inch, header_text)
    
    # Date
    canvas.setFont("Helvetica", 9)
    current_date = datetime.now().strftime("%d/%m/%Y")
    canvas.drawRightString(doc.width + doc.leftMargin, doc.height + doc.topMargin - 0.25*inch, f"Date: {current_date}")
    
    canvas.restoreState()

@main.route('/generate_picklist_pdf')
def generate_picklist_pdf():
    show_further_weeks = request.args.get('show_further_weeks', 'false').lower() == 'true'
    extended_weeks = request.args.getlist('extended_weeks')
    
    # Define default weeks
    default_weeks = ['Arrears', 'Week - 1', 'Week - 2', 'Week - 3', 'Week - 4']
    
    # Add extended weeks if show_further_weeks is True
    if show_further_weeks:
        default_weeks.extend(extended_weeks)

    # Create directory if it doesn't exist
    pdf_dir = os.path.join(current_app.root_path, 'picklist', 'separate_sheets')
    os.makedirs(pdf_dir, exist_ok=True)

    # Query all picklist items with correct sorting
    picklist_query = Picklist.query.order_by(
        Picklist.account_id.asc(),
        case(
            (Picklist.week == 'Arrears', 0),
            (Picklist.week == 'Week - 1', 1),
            (Picklist.week == 'Week - 2', 2),
            (Picklist.week == 'Week - 3', 3),
            (Picklist.week == 'Week - 4', 4),
            (Picklist.week == 'Week - 5', 5),
            (Picklist.week == 'Week - 6', 6),
            (Picklist.week == 'Week - 7', 7),
            (Picklist.week == 'Week - 8', 8),
            else_=9
        ),
        func.date_format(Picklist.required_date, '%d-%m-%Y').asc(),
        Picklist.required_day.asc(),
        Picklist.stock_code.asc()
    )

    # Apply week filter
    picklist_query = picklist_query.filter(Picklist.week.in_(default_weeks))

    picklist_items = picklist_query.all()

    # Group items by account_id, week, and required_date
    grouped_items = {}
    for item in picklist_items:
        if item.account_id not in grouped_items:
            grouped_items[item.account_id] = {}
        if item.week not in grouped_items[item.account_id]:
            grouped_items[item.account_id][item.week] = {}
        if item.required_date not in grouped_items[item.account_id][item.week]:
            grouped_items[item.account_id][item.week][item.required_date] = []
        grouped_items[item.account_id][item.week][item.required_date].append(item)

    # Generate PDF for each account
    pdf_files = []
    for account_id, weeks in grouped_items.items():
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter),
                                leftMargin=0.25*inch, rightMargin=0.25*inch,
                                topMargin=0.75*inch, bottomMargin=0.5*inch)
        elements = []

        class NumberedCanvas(canvas.Canvas):
            def __init__(self, *args, **kwargs):
                canvas.Canvas.__init__(self, *args, **kwargs)
                self._saved_page_states = []

            def showPage(self):
                self._saved_page_states.append(dict(self.__dict__))
                self._startPage()

            def save(self):
                num_pages = len(self._saved_page_states)
                for state in self._saved_page_states:
                    self.__dict__.update(state)
                    self.draw_page_number(num_pages)
                    canvas.Canvas.showPage(self)
                canvas.Canvas.save(self)

            def draw_page_number(self, page_count):
                self.setFont("Helvetica", 9)
                self.drawRightString(
                    self._pagesize[0] - 0.5*inch,
                    0.5*inch,
                    f"Page {self._pageNumber} of {page_count}"
                )

        def header(canvas, doc):
            canvas.saveState()
            logo_path = os.path.join(current_app.root_path, 'static', 'logo', 'ecam_logo.jpg')
            if os.path.exists(logo_path):
                canvas.drawImage(logo_path, doc.leftMargin, doc.height + doc.topMargin - 0.25*inch, width=0.75*inch, height=0.5*inch)
            canvas.setFont("Helvetica-Bold", 14)
            canvas.drawCentredString(doc.width/2.0 + doc.leftMargin, doc.height + doc.topMargin - 0.1*inch, f"ECAM ENGINEERING LIMITED - PICKLIST - {account_id}")
            canvas.setFont("Helvetica", 9)
            current_date = datetime.now().strftime("%d/%m/%Y")
            canvas.drawRightString(doc.width + doc.leftMargin, doc.height + doc.topMargin - 0.1*inch, f"Date: {current_date}")
            canvas.restoreState()

        for week in default_weeks:
            if week in weeks:
                # Add a page break before each new week, except for the first one
                if elements:
                    elements.append(PageBreak())

                for required_date, items in sorted(weeks[week].items(), key=lambda x: datetime.strptime(x[0].strftime('%d-%m-%Y'), '%d-%m-%Y')):
                    # Sort items by stock_code
                    items.sort(key=lambda x: x.stock_code)

                    # Table data
                    data = [['Account', 'Stock Code', 'Issue', 'Required Date', 'Required Day', 'Required Quantity', 'Reference', 'Location', 'Message', 'Week']]
                    for item in items:
                        data.append([
                            item.account_id,
                            item.stock_code,
                            item.issue,
                            item.required_date.strftime('%d-%m-%Y'),
                            item.required_day,
                            item.required_quantity,
                            item.order_reference,
                            item.location,
                            item.message,
                            item.week
                        ])

                    # Create the table
                    table = Table(data, repeatRows=1)
                    style = TableStyle([
                        ('GRID', (0, 0), (-1, -1), 1, colors.black),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                        ('BACKGROUND', (1, 1), (1, -1), colors.lightblue),
                        ('BACKGROUND', (5, 1), (5, -1), colors.lightblue),
                    ])
                    table.setStyle(style)

                    # Check if there's enough space for the table
                    available_height = doc.height - doc.bottomMargin - doc.topMargin
                    table_height = table.wrap(doc.width, available_height)[1]

                    if len(elements) > 0 and table_height > available_height:
                        elements.append(PageBreak())

                    elements.append(KeepTogether(table))
                    elements.append(Spacer(0.15, 0.1 * inch))  # Reduced spacer size

        # Build the PDF
        doc.build(elements, onFirstPage=header, onLaterPages=header, canvasmaker=NumberedCanvas)
        buffer.seek(0)

        # Save the PDF
        pdf_path = os.path.join(pdf_dir, f'picklist_{account_id}.pdf')
        with open(pdf_path, 'wb') as f:
            f.write(buffer.getvalue())
        pdf_files.append(pdf_path)

    # Merge PDFs
    merger = PdfMerger()
    for pdf in pdf_files:
        merger.append(pdf)

    output = BytesIO()
    merger.write(output)
    merger.close()

    # Clean up separate PDF files
    for pdf in pdf_files:
        os.remove(pdf)

    # At the end of the function, replace the existing send_file line with:
    current_datetime = datetime.now()
    current_date = current_datetime.strftime("%d_%m_%Y")
    current_time = current_datetime.strftime("%H_%M")
    filename = f"Picklist_{current_date}_{current_time}.pdf"

    output.seek(0)
    return send_file(output, as_attachment=True, download_name=filename, mimetype='application/pdf')

@main.route('/generate_cancelled_list_pdf')
def generate_cancelled_list_pdf():
    show_further_weeks = request.args.get('show_further_weeks', 'false').lower() == 'true'
    extended_weeks = request.args.getlist('extended_weeks')

    # Define default weeks
    default_weeks = ['Arrears', 'Week - 1', 'Week - 2', 'Week - 3', 'Week - 4']

    # Add extended weeks if show_further_weeks is True
    if show_further_weeks:
        default_weeks.extend(extended_weeks)

    # Create directory if it doesn't exist
    pdf_dir = os.path.join(current_app.root_path, 'cancelled_list', 'separate_sheets')
    os.makedirs(pdf_dir, exist_ok=True)

    # Query all cancelled list items
    cancelled_query = CancelledList.query.order_by(
        CancelledList.account_id.asc(),
        case(
            (CancelledList.week == 'Arrears', 0),
            (CancelledList.week == 'Week - 1', 1),
            (CancelledList.week == 'Week - 2', 2),
            (CancelledList.week == 'Week - 3', 3),
            (CancelledList.week == 'Week - 4', 4),
            (CancelledList.week == 'Week - 5', 5),
            (CancelledList.week == 'Week - 6', 6),
            (CancelledList.week == 'Week - 7', 7),
            (CancelledList.week == 'Week - 8', 8),
            else_=9
        ),
        func.date_format(CancelledList.required_date, '%d-%m-%Y'),
        CancelledList.required_day,
        CancelledList.stock_code
    )

    # Apply week filter
    cancelled_query = cancelled_query.filter(CancelledList.week.in_(default_weeks))

    cancelled_items = cancelled_query.all()

    # Group items by account_id, week, and required_date
    grouped_items = {}
    for item in cancelled_items:
        if item.account_id not in grouped_items:
            grouped_items[item.account_id] = {}
        if item.week not in grouped_items[item.account_id]:
            grouped_items[item.account_id][item.week] = {}
        if item.required_date not in grouped_items[item.account_id][item.week]:
            grouped_items[item.account_id][item.week][item.required_date] = []
        grouped_items[item.account_id][item.week][item.required_date].append(item)

    # Generate PDF for each account
    pdf_files = []
    for account_id, weeks in grouped_items.items():
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter),
                                leftMargin=0.25*inch, rightMargin=0.25*inch,
                                topMargin=0.75*inch, bottomMargin=0.5*inch)
        elements = []

        for week in default_weeks:
            if week in weeks:
                # Add a page break before each new week, except for the first one
                if elements:
                    elements.append(PageBreak())

                for required_date, items in sorted(weeks[week].items(), key=lambda x: datetime.strptime(x[0].strftime('%d-%m-%Y'), '%d-%m-%Y')):
                    # Sort items by stock_code
                    items.sort(key=lambda x: x.stock_code)

                    # Table data
                    data = [['Account', 'Stock Code', 'Issue', 'Required Date', 'Required Day', 'Required Quantity', 'Reference', 'Location', 'Message', 'Week']]
                    for item in items:
                        data.append([
                            item.account_id,
                            item.stock_code,
                            item.issue,
                            item.required_date.strftime('%d-%m-%Y'),
                            item.required_day,
                            item.required_quantity,
                            item.order_reference,
                            item.location,
                            item.message,
                            item.week
                        ])

                    # Create the table
                    table = Table(data, repeatRows=1)
                    style = TableStyle([
                        ('GRID', (0, 0), (-1, -1), 1, colors.black),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('BACKGROUND', (0, 0), (-1, 0), colors.darkgrey),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                        ('BACKGROUND', (1, 1), (1, -1), colors.lightgrey),
                        ('BACKGROUND', (5, 1), (5, -1), colors.lightgrey),
                    ])
                    table.setStyle(style)

                    # Check if there's enough space for the table
                    available_height = doc.height - doc.bottomMargin - doc.topMargin
                    table_height = table.wrap(doc.width, available_height)[1]

                    if len(elements) > 0 and table_height > available_height:
                        elements.append(PageBreak())

                    elements.append(KeepTogether(table))
                    elements.append(Spacer(0.15, 0.1 * inch))  # Reduced spacer size

        # Build the PDF
        doc.build(elements, onFirstPage=lambda canvas, doc: cancelled_header(canvas, doc, account_id),
                  onLaterPages=lambda canvas, doc: cancelled_header(canvas, doc, account_id),
                  canvasmaker=NumberedCanvas)

        buffer.seek(0)

        # Save the PDF
        pdf_path = os.path.join(pdf_dir, f'cancelled_list_{account_id}.pdf')
        with open(pdf_path, 'wb') as f:
            f.write(buffer.getvalue())
        pdf_files.append(pdf_path)

    # Merge PDFs
    merger = PdfMerger()
    for pdf in pdf_files:
        merger.append(pdf)

    output = BytesIO()
    merger.write(output)
    merger.close()

    # Clean up separate PDF files
    for pdf in pdf_files:
        os.remove(pdf)

    # Generate filename with current date and time
    current_datetime = datetime.now()
    current_date = current_datetime.strftime("%d_%m_%Y")
    current_time = current_datetime.strftime("%H_%M")
    filename = f"Cancelled_List_{current_date}_{current_time}.pdf"

    output.seek(0)
    return send_file(output, as_attachment=True, download_name=filename, mimetype='application/pdf')

def cancelled_header(canvas, doc, account_id):
    canvas.saveState()
    styles = getSampleStyleSheet()

    # Logo
    logo_path = os.path.join(current_app.root_path, 'static', 'logo', 'ecam_logo.jpg')
    if os.path.exists(logo_path):
        canvas.drawImage(logo_path, doc.leftMargin, doc.height + doc.topMargin - 0.25*inch, width=0.75*inch, height=0.5*inch)

    # Header text
    header_text = f"ECAM ENGINEERING LIMITED - CANCELLED LIST - {account_id}"
    canvas.setFont("Helvetica-Bold", 14)
    canvas.drawCentredString(doc.width/2.0 + doc.leftMargin, doc.height + doc.topMargin - 0.1*inch, header_text)

    # Date
    canvas.setFont("Helvetica", 9)
    current_date = datetime.now().strftime("%d/%m/%Y")
    canvas.drawRightString(doc.width + doc.leftMargin, doc.height + doc.topMargin - 0.1*inch, f"Date: {current_date}")

    canvas.restoreState()

@main.route('/profile')
@login_required
def profile():
    selected_account_id = request.args.get('account_id', 'All')
    selected_week = request.args.get('week', 'All')
    exclude_cancelled = request.args.get('exclude_cancelled', 'false').lower() == 'true'
    show_price_missing = request.args.get('show_price_missing', 'false').lower() == 'true'

    # Update unit prices from stock_price table
    update_unit_prices()

    week_order = case(
        (SalesOrder.week == 'Arrears', 0),
        *[(SalesOrder.week == f'Week - {i}', i) for i in range(1, 53)],
        else_=53
    )

    sales_query = SalesOrder.query.order_by(
        SalesOrder.account_id.asc(),
        week_order,
        func.date_format(SalesOrder.required_date, '%d-%m-%Y').asc(),
        SalesOrder.stock_code.asc()
    )

    if selected_account_id != 'All':
        sales_query = sales_query.filter(SalesOrder.account_id == selected_account_id)
    if selected_week != 'All':
        sales_query = sales_query.filter(SalesOrder.week == selected_week)
    if exclude_cancelled:
        sales_query = sales_query.filter(SalesOrder.required_quantity != 0)
    if show_price_missing:
        sales_query = sales_query.filter(and_(SalesOrder.unit_price == 0, SalesOrder.required_quantity != 0))

    sales_orders = sales_query.all()

    account_ids = db.session.query(SalesOrder.account_id).distinct().order_by(SalesOrder.account_id).all()
    account_ids = [account_id[0] for account_id in account_ids]

    weeks = db.session.query(SalesOrder.week).distinct().order_by(SalesOrder.week).all()
    weeks = [week[0] for week in weeks]

    return render_template('profile.html', 
                           name=current_user, 
                           sales_orders=sales_orders, 
                           account_ids=account_ids, 
                           weeks=weeks, 
                           selected_account_id=selected_account_id, 
                           selected_week=selected_week, 
                           exclude_cancelled=exclude_cancelled,
                           show_price_missing=show_price_missing)

@main.route('/download_prices_missing_csv')
@login_required
def download_prices_missing_csv():
    # Query for sales orders with unit price 0 and required quantity not 0
    sales_query = SalesOrder.query.filter(
        and_(SalesOrder.unit_price == 0, SalesOrder.required_quantity != 0)
    ).order_by(SalesOrder.stock_code).distinct(SalesOrder.stock_code)

    # Prepare CSV data
    csv_data = StringIO()
    csv_writer = csv.writer(csv_data)

    # Write header
    csv_writer.writerow(['account_id', 'stock_code', 'unit_price'])

    # Write data rows
    for order in sales_query:
        csv_writer.writerow([order.account_id, order.stock_code, order.unit_price])

    # Prepare the response
    csv_data.seek(0)
    return send_file(
        BytesIO(csv_data.getvalue().encode('utf-8')),
        mimetype='text/csv',
        as_attachment=True,
        download_name='PriceMissingList.csv'
    )

@main.route('/generate_sales_workbook')
@login_required
def generate_sales_workbook():
    # Query the sales order data from the database
    sales_orders = SalesOrder.query.all()

    # Create a new workbook
    workbook = Workbook()

    # Define the column headers
    headers = ['Stock Code', 'Issue', 'Required Date', 'Required Quantity', 'Order Reference', 'Location', 'Message', 'Last Delivery Note', 'Last Delivery Date', 'Week', 'Month', 'Unit Price', 'Sale Price']

    # Define the desired account_id order
    account_id_order = ['BAM002', 'BAM003', 'BAM004', 'BAM005', 'BAM007', 'BAM008', 'BAM009', 'BAM011', 'BAM018']

    # Define the desired week order
    week_order = ['Arrears'] + [f'Week - {i}' for i in range(1, 53)]

    # Group sales orders by account_id
    grouped_orders = {}
    for order in sales_orders:
        if order.account_id not in grouped_orders:
            grouped_orders[order.account_id] = []
        grouped_orders[order.account_id].append(order)

    # Create a sheet for each account_id in the desired order
    for account_id in account_id_order:
        if account_id in grouped_orders:
            orders = grouped_orders[account_id]
            sheet = workbook.create_sheet(title=account_id)

        # Write the headers to the sheet
        sheet.append(headers)

        # Set column widths to 18 for all columns
        for column_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
            sheet.column_dimensions[column_letter].width = 15

        # Sort the sales orders by week and required_date
        sorted_orders = sorted(orders, key=lambda x: (week_order.index(x.week) if x.week in week_order else float('inf'), x.required_date))

        # Write the sorted sales order data to the sheet
        for order in sorted_orders:
            row = [
                order.stock_code,
                order.issue,
                order.required_date.strftime('%d-%m-%Y'),  # Change date format here
                order.required_quantity,
                order.order_reference,
                order.location,
                order.message,
                order.last_delivery_note,
                order.last_delivery_date.strftime('%d-%m-%Y') if order.last_delivery_date else '',  # Change date format here
                order.week,
                order.month,
                order.unit_price,
                order.sale_price
            ]
            sheet.append(row)

        # Apply the 'Blue Table Style 9' to the table
        table_range = f"A1:{get_column_letter(len(headers))}{len(sorted_orders)+1}"
        table = Table(displayName=f"Table_{account_id}", ref=table_range)
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style
        sheet.add_table(table) 

        # Apply formatting to the headers
        header_font = Font(bold=True, color="FFFFFF") # Set header font color to white
        for cell in sheet[1]:
            cell.font = header_font

    # Remove the default sheet
    workbook.remove(workbook.active)

    # Save the workbook to a BytesIO object
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    # Generate a filename with the current date and time
    current_datetime = datetime.now()
    current_date = current_datetime.strftime("%d_%m_%Y")
    current_time = current_datetime.strftime("%H_%M")
    filename = f"Sales_Orders_{current_date}_{current_time}.xlsx"

    # Return the Excel file as a downloadable attachment
    return send_file(output, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

def update_unit_prices():
    # Subquery to get the latest unit price for each stock code
    latest_prices = db.session.query(
        StockPrice.stock_code,
        func.max(StockPrice.unit_price).label('latest_price')
    ).group_by(StockPrice.stock_code).subquery()

    # Update SalesOrder table
    update_stmt = SalesOrder.__table__.update().values(
        unit_price=func.coalesce(latest_prices.c.latest_price, 0),
        sale_price=SalesOrder.required_quantity * func.coalesce(latest_prices.c.latest_price, 0)
    ).where(
        SalesOrder.stock_code == latest_prices.c.stock_code
    )

    db.session.execute(update_stmt)

    # Update remaining rows with no matching stock price
    update_remaining_stmt = SalesOrder.__table__.update().values(
        unit_price=0,
        sale_price=0
    ).where(
        SalesOrder.unit_price.is_(None)
    )

    db.session.execute(update_remaining_stmt)
    db.session.commit()


@main.route('/stock_details', methods=['GET', 'POST'])
@login_required
def stock_details():
    if request.method == 'POST':
        if 'file' in request.files:
            file = request.files['file']
            if file.filename == '':
                flash('No file selected', 'error')
                return redirect(url_for('main.stock_details'))
            if not file.filename.lower().endswith('.csv'):
                flash('Invalid file type. Please upload a CSV file.', 'error')
                return redirect(url_for('main.stock_details'))
            try:
                # Read the CSV file
                df = pd.read_csv(file)
                
                # Check if required columns exist
                required_columns = ['account_id', 'stock_code', 'unit_price']
                if not all(col.lower() in map(str.lower, df.columns) for col in required_columns):
                    missing_columns = [col for col in required_columns if col.lower() not in map(str.lower, df.columns)]
                    flash(f'Invalid CSV format. Missing columns: {", ".join(missing_columns)}', 'error')
                    return redirect(url_for('main.stock_details'))
                
                # Normalize column names
                df.columns = df.columns.str.lower()
                
                # Handle NaN values in account_id
                df['account_id'] = df['account_id'].fillna('Unknown')
                
                # Process the file
                added_count = 0
                updated_count = 0
                for _, row in df.iterrows():
                    try:
                        account_id = row['account_id']
                        stock_code = row['stock_code']
                        unit_price = float(row['unit_price'])
                        
                        # Always create a new StockPrice entry
                        stock_price = StockPrice(account_id=account_id, stock_code=stock_code, unit_price=unit_price)
                        db.session.add(stock_price)
                        added_count += 1
                        
                    except ValueError as ve:
                        flash(f'Error processing row: {row}. Error: {str(ve)}', 'error')
                    except Exception as e:
                        flash(f'Unexpected error processing row: {row}. Error: {str(e)}', 'error')
                
                db.session.commit()
                flash(f'Successfully added {added_count} stock prices', 'success')
            except Exception as e:
                flash(f'An unexpected error occurred. Invalid CSV File. Error: {str(e)}', 'error')
            
            return redirect(url_for('main.stock_details'))
        
        elif 'account_id' in request.form and 'stock_code' in request.form and 'unit_price' in request.form:
            # Handle single stock price update
            account_id = request.form['account_id']
            stock_code = request.form['stock_code']
            unit_price = request.form['unit_price']
            try:
                update_stock_price(account_id, stock_code, float(unit_price))
                db.session.commit()
                flash(f'Successfully updated price for {account_id} - {stock_code}', 'success')
            except ValueError:
                flash('Invalid unit price. Please enter a valid number.', 'error')
            except Exception as e:
                flash(f'An error occurred: {str(e)}', 'error')
            return redirect(url_for('main.stock_details'))

    # GET request handling
    stock_prices = StockPrice.query.order_by(StockPrice.account_id, StockPrice.stock_code).all()
    
    # Group stock prices by account_id and stock_code
    grouped_stock_prices = []
    for (account_id, stock_code), group in groupby(stock_prices, key=lambda x: (x.account_id, x.stock_code)):
        group_list = list(group)
        if len(group_list) > 1 and len(set(item.unit_price for item in group_list)) > 1:
            grouped_stock_prices.append(group_list)
    
    return render_template('stock_details.html', stock_prices=stock_prices, grouped_stock_prices=grouped_stock_prices)

@main.route('/update_selected_prices', methods=['POST'])
@login_required
def update_selected_prices():
    account_id = request.form['account_id']
    stock_code = request.form['stock_code']
    unit_price = request.form['unit_price']

    try:
        update_stock_price(account_id, stock_code, float(unit_price))
        db.session.commit()
        flash(f'Successfully updated price for {account_id} - {stock_code}', 'success')
    except ValueError:
        flash('Invalid unit price. Please enter a valid number.', 'error')
    except Exception as e:
        db.session.rollback()
        flash(f'An error occurred: {str(e)}', 'error')

    return redirect(url_for('main.stock_details'))

def update_stock_price(account_id, stock_code, unit_price):
    stock_price = StockPrice.query.filter_by(account_id=account_id, stock_code=stock_code).first()
    if stock_price:
        stock_price.unit_price = unit_price
        stock_price.last_updated = datetime.now()
    else:
        stock_price = StockPrice(account_id=account_id, stock_code=stock_code, unit_price=unit_price)
        db.session.add(stock_price)

    # Update the related SalesOrder table
    SalesOrder.query.filter_by(account_id=account_id, stock_code=stock_code).update(
        {SalesOrder.unit_price: unit_price, SalesOrder.sale_price: SalesOrder.required_quantity * unit_price},
        synchronize_session=False
    )

    # Commit the changes to ensure SalesOrder is updated
    db.session.commit()

    # Update forecasts and related tables
    with db.engine.connect() as connection:
        # Call the generate_forecasts stored procedure
        connection.execute(text("CALL generate_forecasts()"))

    # Commit the changes again after updating forecasts
    db.session.commit()

@main.route('/remove_duplicates', methods=['POST'])
@login_required
def remove_duplicates():
    data = request.json
    ids_to_remove = data.get('ids', [])

    try:
        StockPrice.query.filter(StockPrice.id.in_(ids_to_remove)).delete(synchronize_session=False)
        db.session.commit()
        return jsonify({'status': 'success', 'message': f'Successfully removed {len(ids_to_remove)} item(s)'}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'status': 'error', 'message': f'An error occurred: {str(e)}'}), 500
    
@main.route('/remove_stock_prices', methods=['POST'])
@login_required
def remove_stock_prices():
    data = request.json
    ids_to_remove = data.get('ids', [])

    try:
        StockPrice.query.filter(StockPrice.id.in_(ids_to_remove)).delete(synchronize_session=False)
        db.session.commit()
        return jsonify({'status': 'success', 'message': f'Successfully removed {len(ids_to_remove)} stock price(s)'}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'status': 'error', 'message': f'An error occurred: {str(e)}'}), 500
    
@main.route('/update_stock_prices', methods=['POST'])
@login_required
def update_stock_prices():
    data = request.json
    stocks_to_update = data.get('stocks', [])
    updated_count = 0
    try:
        for stock in stocks_to_update:
            stock_price = StockPrice.query.get(stock['id'])
            if stock_price:
                if (stock_price.account_id != stock['account_id'] or
                    stock_price.stock_code != stock['stock_code'] or
                    abs(float(stock_price.unit_price) - float(stock['unit_price'])) > 0.001):
                    stock_price.account_id = stock['account_id']
                    stock_price.stock_code = stock['stock_code']
                    stock_price.unit_price = stock['unit_price']
                    stock_price.last_updated = datetime.now()
                    updated_count += 1
        
        if updated_count > 0:
            db.session.commit()
            # Update related tables
            update_related_tables()
            return jsonify({'status': 'success', 'message': f'Successfully updated {updated_count} stock price(s)'}), 200
        else:
            return jsonify({'status': 'info', 'message': 'No changes were made'}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'status': 'error', 'message': f'An error occurred: {str(e)}'}), 500

def update_related_tables():
    # Update the SalesOrder table
    db.session.execute(text("""
        UPDATE sales_order so
        JOIN stock_price sp ON so.stock_code = sp.stock_code
        SET so.unit_price = sp.unit_price,
            so.sale_price = so.required_quantity * sp.unit_price
    """))
    
    # Call stored procedures to update other tables
    db.session.execute(text("CALL update_week_column()"))
    db.session.execute(text("CALL create_picklist()"))
    db.session.execute(text("CALL create_cancelled_list()"))
    db.session.execute(text("CALL generate_forecasts()"))
    db.session.commit()

@main.route('/clear_stock_prices', methods=['POST'])
@login_required
def clear_stock_prices():
    try:
        num_deleted = db.session.query(StockPrice).delete()
        db.session.commit()
        flash(f'Successfully deleted {num_deleted} stock prices', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'An error occurred while clearing stock prices: {str(e)}', 'error')
    return redirect(url_for('main.stock_details'))

def get_duplicates_with_different_prices():
    subquery = db.session.query(
        StockPrice.account_id,
        StockPrice.stock_code,
        func.count(StockPrice.id).label('count'),
        func.count(func.distinct(StockPrice.unit_price)).label('price_count')
    ).group_by(StockPrice.account_id, StockPrice.stock_code).subquery()

    duplicates = db.session.query(StockPrice).join(
        subquery,
        and_(
            StockPrice.account_id == subquery.c.account_id,
            StockPrice.stock_code == subquery.c.stock_code
        )
    ).filter(
        subquery.c.count > 1,
        subquery.c.price_count > 1
    ).order_by(StockPrice.account_id, StockPrice.stock_code, StockPrice.unit_price).all()

    return duplicates

@main.route('/download_duplicate_list')
@login_required
def download_duplicate_list():
    duplicates = get_duplicates_with_different_prices()
    
    # Create a CSV string
    output = StringIO()
    writer = csv.writer(output)
    
    # Write header
    writer.writerow(['account_id', 'stock_code', 'unit_price'])
    
    # Write data
    for price in duplicates:
        writer.writerow([price.account_id, price.stock_code, price.unit_price])
    
    # Create the response
    response = make_response(output.getvalue())
    response.headers["Content-Disposition"] = "attachment; filename=duplicate_stock_list.csv"
    response.headers["Content-type"] = "text/csv"
    
    return response

@main.route('/download_csv_format')
@login_required
def download_csv_format():
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(['account_id', 'stock_code', 'unit_price'])
    
    response = make_response(output.getvalue())
    response.headers["Content-Disposition"] = "attachment; filename=stock_price_template.csv"
    response.headers["Content-type"] = "text/csv"
    return response

@main.route('/generate_stock_price_workbook')
@login_required
def generate_stock_price_workbook():
    # Query the stock price data from the database
    stock_prices = StockPrice.query.order_by(StockPrice.account_id, StockPrice.stock_code).all()

    # Create a new workbook and select the active sheet
    workbook = Workbook()
    sheet = workbook.active

    # Define the column headers
    headers = ['Account ID', 'Stock Code', 'Unit Price', 'Last Updated']

    # Write the headers to the sheet
    sheet.append(headers)

    # Write the stock price data to the sheet
    for price in stock_prices:
        row = [
            price.account_id,
            price.stock_code,
            price.unit_price,
            price.last_updated.strftime('%Y-%m-%d %H:%M:%S %Z')
        ]
        sheet.append(row)

    # Set column widths to 18 for all columns
    for column_letter in ['A', 'B', 'C', 'D']:
        sheet.column_dimensions[column_letter].width = 18

    # Apply table formatting
    table_range = f"A1:D{len(stock_prices)+1}"
    table = Table(displayName="StockPricesTable", ref=table_range)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    sheet.add_table(table)

    # Apply formatting to the headers
    header_font = Font(bold=True, color="FFFFFF")
    for cell in sheet[1]:
        cell.font = header_font

    # Save the workbook to a BytesIO object
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    # Generate a filename with the current date and time
    current_datetime = datetime.now()
    current_date = current_datetime.strftime("%d_%m_%Y")
    current_time = current_datetime.strftime("%H_%M")
    filename = f"Stock_Prices_{current_date}_{current_time}.xlsx"

    # Return the Excel file as a downloadable attachment
    return send_file(output, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@main.route('/generate_stock_price_csv')
@login_required
def generate_stock_price_csv():
    # Query the stock price data from the database
    stock_prices = StockPrice.query.order_by(StockPrice.account_id, StockPrice.stock_code).all()

    # Create a StringIO object to write the CSV data
    csv_data = StringIO()
    writer = csv.writer(csv_data)

    # Write the header row
    writer.writerow(['account_id', 'stock_code', 'unit_price'])

    # Write the stock price data rows
    for price in stock_prices:
        writer.writerow([price.account_id, price.stock_code, price.unit_price])

    # Generate a filename with the current date and time
    current_datetime = datetime.now()
    current_date = current_datetime.strftime("%d_%m_%Y")
    current_time = current_datetime.strftime("%H_%M")
    filename = f"Stock_Prices_{current_date}_{current_time}.csv"

    # Create a response object with the CSV data
    response = make_response(csv_data.getvalue())
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"
    response.headers["Content-type"] = "text/csv"

    return response

def get_weekly_forecast(page):
    weeks_per_page = 9
    total_weeks = 52
    start_week = (page - 1) * weeks_per_page + 1
    end_week = start_week + weeks_per_page - 1

    week_columns = [f"Week - {i}" for i in range(start_week, min(end_week + 1, total_weeks + 1))]

    weekly_forecasts = db.session.query(
        WeeklyForecasts.account_id.label("Account"),
        WeeklyForecasts.Arrears,
        *[getattr(WeeklyForecasts, f"Week - {i}").label(f"Week - {i}") for i in range(start_week, min(end_week + 1, total_weeks + 1))],
        WeeklyTotals.Arrears.label("Total_Arrears"),
        WeeklyTotals.Weeks_1_9.label("Total_Weeks_1-9"),
        WeeklyTotals.Weeks_10_18.label("Total_Weeks_10-18"),
        WeeklyTotals.Weeks_19_27.label("Total_Weeks_19-27"),
        WeeklyTotals.Weeks_28_36.label("Total_Weeks_28-36"),
        WeeklyTotals.Weeks_37_45.label("Total_Weeks_37-45"),
        WeeklyTotals.Weeks_46_52.label("Total_Weeks_46-52"),
        GrandTotal.weekly_arrears,
        GrandTotal.weekly.label("Grand_Total")
    ).join(WeeklyTotals, WeeklyForecasts.account_id == WeeklyTotals.account_id
    ).join(GrandTotal, WeeklyForecasts.account_id == GrandTotal.account_id
    ).order_by(WeeklyForecasts.account_id).all()

    weekly_forecasts_dict = []
    for forecast in weekly_forecasts:
        forecast_dict = {
            'Account': forecast.Account,
            'Arrears': forecast.Arrears,
            **{week: getattr(forecast, week) for week in week_columns},
            'Total': getattr(forecast, f"Total_Weeks_{start_week}-{min(end_week, 52)}"),
            'Grand Total': forecast.Grand_Total
        }
        weekly_forecasts_dict.append(forecast_dict)

    weekly_total_row = {
        'Account': 'Total',
        'Arrears': sum(forecast['Arrears'] for forecast in weekly_forecasts_dict if forecast['Arrears'] is not None),
        **{week: sum(forecast[week] for forecast in weekly_forecasts_dict if forecast[week] is not None) for week in week_columns},
        'Total': sum(forecast['Total'] for forecast in weekly_forecasts_dict if forecast['Total'] is not None),
        'Grand Total': sum(forecast['Grand Total'] for forecast in weekly_forecasts_dict if forecast['Grand Total'] is not None)
    }
    weekly_forecasts_dict.append(weekly_total_row)

    total_weekly_pages = (total_weeks + weeks_per_page - 1) // weeks_per_page

    return weekly_forecasts_dict, week_columns, total_weekly_pages

def get_monthly_forecast(page):
    months_per_page = 6
    total_months = 12
    start_month = (page - 1) * months_per_page + 1
    end_month = start_month + months_per_page - 1

    monthly_forecasts = db.session.query(
        MonthlyForecasts.account_id.label("Account"),
        MonthlyForecasts.Arrears,
        *[getattr(MonthlyForecasts, f"Month{i}").label(f"Month{i}") for i in range(1, 13)],
        *[getattr(MonthlyForecasts, f"Month{i}_Value").label(f"Month{i}_Value") for i in range(1, 13)],
        MonthlyTotals.Arrears.label("Total_Arrears"),
        MonthlyTotals.Months_1_6.label("Total_Months_1_6"),
        MonthlyTotals.Months_7_12.label("Total_Months_7_12"),
        GrandTotal.monthly_arrears,
        GrandTotal.monthly.label("Grand_Total")
    ).join(MonthlyTotals, MonthlyForecasts.account_id == MonthlyTotals.account_id
    ).join(GrandTotal, MonthlyForecasts.account_id == GrandTotal.account_id
    ).order_by(MonthlyForecasts.account_id).all()

    monthly_forecasts_dict = []
    month_columns = []
    for forecast in monthly_forecasts:
        forecast_dict = {
            'Account': forecast.Account,
            'Arrears': forecast.Arrears
        }
        for i in range(start_month, min(end_month + 1, total_months + 1)):
            month_name = getattr(forecast, f"Month{i}")
            month_value = getattr(forecast, f"Month{i}_Value")
            forecast_dict[month_name] = month_value
            if month_name not in month_columns:
                month_columns.append(month_name)
        
        if page == 1:
            forecast_dict['Total'] = forecast.Total_Months_1_6
        else:
            forecast_dict['Total'] = forecast.Total_Months_7_12
        
        forecast_dict['Grand Total'] = forecast.Grand_Total
        monthly_forecasts_dict.append(forecast_dict)

    monthly_total_row = {
        'Account': 'Total',
        'Arrears': sum(forecast['Arrears'] for forecast in monthly_forecasts_dict if forecast['Arrears'] is not None),
        **{month: sum(forecast.get(month, 0) for forecast in monthly_forecasts_dict if forecast.get(month) is not None) for month in month_columns},
        'Total': sum(forecast['Total'] for forecast in monthly_forecasts_dict if forecast['Total'] is not None),
        'Grand Total': sum(forecast['Grand Total'] for forecast in monthly_forecasts_dict if forecast['Grand Total'] is not None)
    }
    monthly_forecasts_dict.append(monthly_total_row)

    total_monthly_pages = (total_months + months_per_page - 1) // months_per_page

    return monthly_forecasts_dict, month_columns, total_monthly_pages

@main.route('/sales_forecasts')
@login_required
def sales_forecasts():
    weekly_page = request.args.get('weekly_page', 1, type=int)
    monthly_page = request.args.get('monthly_page', 1, type=int)
    
    weeks_per_page = 9
    months_per_page = 6
    total_weeks = 52
    total_months = 12

    weekly_forecasts, week_columns, total_weekly_pages = get_weekly_forecast(weekly_page)
    monthly_forecasts, month_columns, total_monthly_pages = get_monthly_forecast(monthly_page)

    return render_template('sales_forecasts.html',
                           weekly_forecasts=weekly_forecasts,
                           monthly_forecasts=monthly_forecasts,
                           week_columns=week_columns,
                           month_columns=month_columns,
                           weekly_page=weekly_page,
                           monthly_page=monthly_page,
                           total_weekly_pages=total_weekly_pages,
                           total_monthly_pages=total_monthly_pages,
                           weeks_per_page=weeks_per_page,
                           months_per_page=months_per_page,
                           total_weeks=total_weeks,
                           total_months=total_months)

def header_weekly(canvas, doc):
    logo_path = os.path.join(current_app.root_path, 'static', 'logo', 'ecam_logo.jpg')
    canvas.saveState()
    if os.path.exists(logo_path):
        canvas.drawImage(logo_path, doc.leftMargin, doc.height + doc.topMargin - 0.65*inch, width=1*inch, height=0.65*inch)
    else:
        print(f"Logo file not found at: {logo_path}")
    
    canvas.setFont("Helvetica-Bold", 14)
    canvas.drawCentredString(doc.width/2.0 + doc.leftMargin, doc.height + doc.topMargin - 0.25*inch, "ECAM Engineering Limited - Weekly Forecasts")
    
    canvas.setFont("Helvetica", 10)
    current_date = datetime.now().strftime("%d/%m/%Y")
    canvas.drawRightString(doc.width + doc.leftMargin, doc.height + doc.topMargin - 0.25*inch, f"Created Date: {current_date}")
    
    canvas.restoreState()

def header_monthly(canvas, doc):
    logo_path = os.path.join(current_app.root_path, 'static', 'logo', 'ecam_logo.jpg')
    canvas.saveState()
    if os.path.exists(logo_path):
        canvas.drawImage(logo_path, doc.leftMargin, doc.height + doc.topMargin - 0.65*inch, width=1*inch, height=0.65*inch)
    else:
        print(f"Logo file not found at: {logo_path}")
    
    canvas.setFont("Helvetica-Bold", 14)
    canvas.drawCentredString(doc.width/2.0 + doc.leftMargin, doc.height + doc.topMargin - 0.25*inch, "ECAM Engineering Limited - Monthly Forecasts")
    
    canvas.setFont("Helvetica", 10)
    current_date = datetime.now().strftime("%d/%m/%Y")
    canvas.drawRightString(doc.width + doc.leftMargin, doc.height + doc.topMargin - 0.25*inch, f"Created Date: {current_date}")
    
    canvas.restoreState()

@main.route('/generate_weekly_forecast_pdf')
@login_required
def generate_weekly_forecast_pdf():
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), leftMargin=0.5*inch, rightMargin=0.5*inch, topMargin=1.25*inch, bottomMargin=0.5*inch)
    elements = []

    # Add a spacer after the header
    elements.append(Spacer(1, 0.25*inch))

    styles = getSampleStyleSheet()
    title_style = styles['Heading1']
    title_style.alignment = 1  # Center alignment

    total_weekly_pages = (52 + 8) // 9

    for page in range(1, total_weekly_pages + 1):
        weekly_forecasts, week_columns, _ = get_weekly_forecast(page)

        if page == 1:
            data = [['Account', 'Arrears'] + week_columns + ['Total', 'Grand Total']]
        else:
            data = [['Account'] + week_columns + ['Total', 'Grand Total']]

        for forecast in weekly_forecasts:
            if page == 1:
                row = [forecast['Account'], f"{forecast['Arrears']:.2f}"]
            else:
                row = [forecast['Account']]
            row.extend([f"{forecast.get(week, 0):.2f}" for week in week_columns])
            row.extend([f"{forecast['Total']:.2f}", f"{forecast['Grand Total']:.2f}"])
            data.append(row)

        table = Table(data)
        style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('BACKGROUND', (0, 1), (0, -1), colors.darkblue),
            ('TEXTCOLOR', (0, 1), (0, -1), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('BACKGROUND', (1, 1), (-2, -2), colors.white),
            ('TEXTCOLOR', (1, 1), (-2, -2), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (1, 1), (-2, -2), 'Helvetica'),
            ('FONTSIZE', (1, 1), (-2, -2), 10),
            ('BACKGROUND', (1, -1), (-2, -1), colors.lightyellow),
            ('TEXTCOLOR', (1, -1), (-2, -1), colors.black),
            ('FONTNAME', (1, -1), (-2, -1), 'Helvetica-Bold'),
            ('BACKGROUND', (-2, 1), (-2, -1), colors.lightyellow),
            ('TEXTCOLOR', (-2, 1), (-2, -1), colors.black),
            ('FONTNAME', (-2, 1), (-2, -1), 'Helvetica-Bold'),
            ('BACKGROUND', (-1, 1), (-1, -1), colors.lightgreen),
            ('TEXTCOLOR', (-1, 1), (-1, -1), colors.black),
            ('FONTNAME', (-1, 1), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ])
        table.setStyle(style)

        elements.append(KeepTogether(table))
        elements.append(Spacer(1, 0.25 * inch))

    doc.build(elements, onFirstPage=header_weekly, onLaterPages=header_weekly, canvasmaker=NumberedCanvas)
    buffer.seek(0)

    # Generate filename with current date and time
    current_datetime = datetime.now()
    current_date = current_datetime.strftime("%d_%m_%Y")
    current_time = current_datetime.strftime("%H_%M")
    filename = f"Weekly_Forecast_{current_date}_{current_time}.pdf"

    return send_file(buffer, as_attachment=True, download_name=filename, mimetype='application/pdf')

@main.route('/generate_monthly_forecast_pdf')
@login_required
def generate_monthly_forecast_pdf():
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), 
                            leftMargin=0.5*inch, rightMargin=0.5*inch, 
                            topMargin=1.25*inch, bottomMargin=0.5*inch)
    elements = []
    # Add a spacer after the header
    elements.append(Spacer(1, 0.25*inch))

    styles = getSampleStyleSheet()
    title_style = styles['Heading1']
    title_style.alignment = 1  # Center alignment

    total_monthly_pages = 2  # Assuming there are two pages for monthly forecasts
    for page in range(1, total_monthly_pages + 1):
        monthly_forecasts, month_columns, _ = get_monthly_forecast(page)
        
        if page == 1:
            data = [['Account', 'Arrears'] + month_columns + ['Total', 'Grand Total']]
        else:
            data = [['Account'] + month_columns + ['Total', 'Grand Total']]
        
        for forecast in monthly_forecasts:
            if page == 1:
                row = [forecast['Account'], f"{forecast['Arrears']:.2f}"]
            else:
                row = [forecast['Account']]
            row.extend([f"{forecast.get(month, 0):.2f}" for month in month_columns])
            row.extend([f"{forecast['Total']:.2f}", f"{forecast['Grand Total']:.2f}"])
            data.append(row)

        table = Table(data)
        style = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('BACKGROUND', (0, 1), (0, -1), colors.darkblue),
            ('TEXTCOLOR', (0, 1), (0, -1), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('BACKGROUND', (1, 1), (-2, -2), colors.white),
            ('TEXTCOLOR', (1, 1), (-2, -2), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (1, 1), (-2, -2), 'Helvetica'),
            ('FONTSIZE', (1, 1), (-2, -2), 10),
            ('BACKGROUND', (1, -1), (-2, -1), colors.lightyellow),
            ('TEXTCOLOR', (1, -1), (-2, -1), colors.black),
            ('FONTNAME', (1, -1), (-2, -1), 'Helvetica-Bold'),
            ('BACKGROUND', (-2, 1), (-2, -1), colors.lightyellow),
            ('TEXTCOLOR', (-2, 1), (-2, -1), colors.black),
            ('FONTNAME', (-2, 1), (-2, -1), 'Helvetica-Bold'),
            ('BACKGROUND', (-1, 1), (-1, -1), colors.lightgreen),
            ('TEXTCOLOR', (-1, 1), (-1, -1), colors.black),
            ('FONTNAME', (-1, 1), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ]
        table.setStyle(TableStyle(style))
        elements.append(KeepTogether(table))
        elements.append(Spacer(1, 0.25 * inch))

    doc.build(elements, onFirstPage=header_monthly, onLaterPages=header_monthly)
    buffer.seek(0)

    # At the end of the function, replace the existing send_file line with:
    current_datetime = datetime.now()
    current_date = current_datetime.strftime("%d_%m_%Y")
    current_time = current_datetime.strftime("%H_%M")
    filename = f"Monthly_Forecast_{current_date}_{current_time}.pdf"

    return send_file(buffer, as_attachment=True, download_name=filename, mimetype='application/pdf')
@main.route('/generate_sales_forecasts_workbook')
@login_required
def generate_sales_forecasts_workbook():
    # Query the sales forecasts data from the database
    weekly_forecasts = db.session.query(
        WeeklyForecasts.account_id,
        WeeklyForecasts.Arrears,
        *[getattr(WeeklyForecasts, f'Week - {i}') for i in range(1, 53)],
        GrandTotal.weekly.label('Grand_Total')
    ).join(GrandTotal, WeeklyForecasts.account_id == GrandTotal.account_id).all()

    monthly_forecasts = db.session.query(
        MonthlyForecasts.account_id,
        MonthlyForecasts.Arrears,
        *[getattr(MonthlyForecasts, f'Month{i}') for i in range(1, 13)],
        *[getattr(MonthlyForecasts, f'Month{i}_Value') for i in range(1, 13)],
        GrandTotal.monthly.label('Grand_Total')
    ).join(GrandTotal, MonthlyForecasts.account_id == GrandTotal.account_id).all()

    # Create a new workbook
    workbook = Workbook()

    # Create the "Weekly Forecasts" sheet
    weekly_sheet = workbook.create_sheet(title="Weekly Forecasts")

    # Define the column headers for weekly forecasts
    weekly_headers = ['Account', 'Arrears'] + [f'Week - {i}' for i in range(1, 53)] + ['Grand Total']

    # Write the headers to the weekly sheet
    weekly_sheet.append(weekly_headers)

    # Write the weekly forecasts data to the sheet
    for forecast in weekly_forecasts:
        row = [
            forecast.account_id,
            forecast.Arrears,
        ] + [getattr(forecast, f'Week - {i}') for i in range(1, 53)] + [
            forecast.Grand_Total
        ]
        weekly_sheet.append(row)

    # Calculate the total row for weekly forecasts
    total_row = ['Total'] + [
        sum(getattr(forecast, column) for forecast in weekly_forecasts if getattr(forecast, column) is not None)
        for column in weekly_headers[1:-1]
    ] + [sum(forecast.Grand_Total for forecast in weekly_forecasts if forecast.Grand_Total is not None)]
    weekly_sheet.append(total_row)

    # Set column widths to 18 for all columns in the weekly sheet
    for column_letter in ['A', 'B'] + [get_column_letter(i) for i in range(3, 56)]:
        weekly_sheet.column_dimensions[column_letter].width = 18

    # Apply the 'Blue Table Style 9' to the "Weekly Forecasts" table
    weekly_table_range = f"A1:{get_column_letter(len(weekly_headers))}{len(weekly_forecasts)+2}"
    weekly_table = Table(displayName="WeeklyForecastsTable", ref=weekly_table_range)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    weekly_table.tableStyleInfo = style
    weekly_sheet.add_table(weekly_table)

    # Apply bold formatting to the 'Total' row in the "Weekly Forecasts" sheet
    for cell in weekly_sheet[len(weekly_forecasts)+2]:
        cell.font = Font(bold=True)

    # Apply 'Calculation' style to the entire last column in the "Weekly Forecasts" sheet
    grand_total_column = get_column_letter(len(weekly_headers))
    for cell in weekly_sheet[grand_total_column]:
        cell.style = 'Calculation'

    # Apply 'Calculation' style to the entire last row in the "Weekly Forecasts" sheet
    for cell in weekly_sheet[len(weekly_forecasts)+2]:
        cell.style = 'Calculation'


    # Create the "Monthly Forecasts" sheet
    monthly_sheet = workbook.create_sheet(title="Monthly Forecasts")

    # Define the column headers for monthly forecasts
    monthly_headers = ['Account', 'Arrears'] + [getattr(monthly_forecasts[0], f'Month{i}') for i in range(1, 13)] + ['Grand Total']

    # Write the headers to the monthly sheet
    monthly_sheet.append(monthly_headers)

    # Write the monthly forecasts data to the sheet
    for forecast in monthly_forecasts:
        row = [
            forecast.account_id,
            forecast.Arrears,
        ] + [getattr(forecast, f'Month{i}_Value') for i in range(1, 13)] + [
            forecast.Grand_Total
        ]
        monthly_sheet.append(row)

    # Calculate the total row for monthly forecasts
    total_row = ['Total'] + [
        sum(forecast.Arrears for forecast in monthly_forecasts if forecast.Arrears is not None)
    ] + [
        sum(getattr(forecast, f'Month{i}_Value') for forecast in monthly_forecasts if getattr(forecast, f'Month{i}_Value') is not None)
        for i in range(1, 13)
    ] + [
        sum(forecast.Grand_Total for forecast in monthly_forecasts if forecast.Grand_Total is not None)
    ]
    monthly_sheet.append(total_row)

    # Set column widths to 18 for all columns in the monthly sheet
    for column_letter in ['A', 'B'] + [get_column_letter(i) for i in range(3, 16)]:
        monthly_sheet.column_dimensions[column_letter].width = 18

    # Apply the 'Blue Table Style 9' to the "Monthly Forecasts" table
    monthly_table_range = f"A1:{get_column_letter(len(monthly_headers))}{len(monthly_forecasts)+2}"
    monthly_table = Table(displayName="MonthlyForecastsTable", ref=monthly_table_range)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    monthly_table.tableStyleInfo = style
    monthly_sheet.add_table(monthly_table)

    # Apply bold formatting to the 'Total' row in the "Monthly Forecasts" sheet
    for cell in monthly_sheet[len(monthly_forecasts)+2]:
        cell.font = Font(bold=True)

    # Apply 'Calculation' style to the entire last column in the "Monthly Forecasts" sheet
    grand_total_column = get_column_letter(len(monthly_headers))
    for cell in monthly_sheet[grand_total_column]:
        cell.style = 'Calculation'

    # Apply 'Calculation' style to the entire last row in the "Monthly Forecasts" sheet
    for cell in monthly_sheet[len(monthly_forecasts)+2]:
        cell.style = 'Calculation'

    # Remove the default sheet
    workbook.remove(workbook.active)

    # Save the workbook to a BytesIO object
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    # Generate a filename with the current date and time
    current_datetime = datetime.now()
    current_date = current_datetime.strftime("%d_%m_%Y")
    current_time = current_datetime.strftime("%H_%M")
    filename = f"Sales_Forecasts_{current_date}_{current_time}.xlsx"

    # Return the Excel file as a downloadable attachment
    return send_file(output, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@main.route('/clear_app_data', methods=['POST'])
@login_required
def clear_app_data():
    try:
        delete_sales_orders_and_picklist()
        return jsonify({'status': 'success', 'message': 'App data cleared successfully'}), 200
    except Exception as e:
        return jsonify({'status': 'error', 'message': f'Error clearing app data: {str(e)}'}), 500
    
@contextmanager
def session_scope():
    """Provide a transactional scope around a series of operations."""
    session = Session(db.engine)
    try:
        yield session
        session.commit()
    except Exception as e:
        session.rollback()
        raise e
    finally:
        session.close()

@main.route('/hse_uploader', methods=['GET', 'POST'])
@login_required
def hse_uploader():
    if request.method == 'POST':
        try: 
            try:
                # Delete all files in the UPLOAD_FOLDER
                for filename in os.listdir(UPLOAD_FOLDER):
                    file_path = os.path.join(UPLOAD_FOLDER, filename)
                    if os.path.isfile(file_path):
                        os.unlink(file_path)
                    elif os.path.isdir(file_path):
                        shutil.rmtree(file_path)
                logging.info("All previous files deleted successfully.")
            except Exception as e:
                logging.error(f"Error deleting previous files: {str(e)}")
                return jsonify({'status': 'error', 'message': f"Error deleting previous files: {str(e)}"})

            hse_files = request.files.getlist('hse_files')
            if not hse_files:
                return jsonify({'status': 'error', 'message': "No files were uploaded."})

            file_count = 0
            processed_files = []
            error_files = []

            for hse_file in hse_files:
                if hse_file.filename.lower().endswith('.hse'):
                    try:
                        # Save the uploaded file to the specified location
                        file_path = os.path.join(UPLOAD_FOLDER, hse_file.filename)
                        hse_file.save(file_path)

                        logging.info(f"Processing file: {hse_file.filename}")

                        # Process the uploaded file
                        df = read_hse_file(file_path)
                        if df.empty:
                            raise ValueError("The file is empty or couldn't be parsed correctly.")

                        account_id = get_account_id(hse_file.filename)
                        if account_id == 'Unknown':
                            raise ValueError("Unable to determine account ID from filename.")

                        logging.info(f"Account ID: {account_id}")
                        logging.info(f"DataFrame shape: {df.shape}")

                        # Start a new transaction for each file
                        with session_scope() as session:
                            for _, row in df.iterrows():
                                existing_order = session.query(SalesOrder).filter_by(
                                    account_id=account_id,
                                    stock_code=row['Stock Code'],
                                    issue=row['Issue'],
                                    required_date=row['Required Date'],
                                    required_quantity=row['Required Quantities']
                                ).first()

                                if not existing_order:
                                    sales_order = SalesOrder(
                                        account_id=account_id,
                                        stock_code=row['Stock Code'],
                                        issue=row['Issue'],
                                        required_date=row['Required Date'],
                                        required_quantity=row['Required Quantities'],
                                        order_reference=row['Order Reference'],
                                        location=row['Location'],
                                        message=row['Message'],
                                        last_delivery_note=row['Last Delivery Note'] if row['Last Delivery Note'] != 'NO PREVIOUS' else None,
                                        last_delivery_date=row['Last Delivery Date'],
                                        month=row['Month']
                                    )
                                    session.add(sales_order)
                                    file_count += 1

                            # Update unit prices and sale prices
                            stock_prices = session.query(StockPrice).all()
                            for stock_price in stock_prices:
                                session.query(SalesOrder).filter_by(stock_code=stock_price.stock_code).update({
                                    SalesOrder.unit_price: stock_price.unit_price,
                                    SalesOrder.sale_price: SalesOrder.required_quantity * stock_price.unit_price
                                }, synchronize_session=False)

                            # Execute stored procedures
                            session.execute(text("CALL update_week_column()"))
                            session.execute(text("CALL create_picklist()"))
                            session.execute(text("CALL create_cancelled_list()"))
                            session.execute(text("CALL generate_forecasts()"))                         

                        processed_files.append(hse_file.filename)
                        file_count += 1

                    except Exception as e:
                        error_message = f"Error processing file {hse_file.filename}: {str(e)}"
                        logging.error(error_message)
                        error_files.append((hse_file.filename, str(e)))
                else:
                    error_files.append((hse_file.filename, "Not an HSE file"))

            if file_count > 0:
                message = f"{file_count} new sales orders have been added and processed successfully!"
                if error_files:
                    message += f" However, there were issues with {len(error_files)} file(s)."
                return jsonify({'status': 'success', 'message': message, 'processed_files': processed_files, 'error_files': error_files})
            elif error_files:
                return jsonify({'status': 'error', 'message': f"There were issues processing {len(error_files)} file(s). Please check the error details.", 'error_files': error_files})
            else:
                return jsonify({'status': 'info', 'message': "No new sales orders were added. The files may have already been processed."})
            
        except Exception as e:
            logging.error(f"Unexpected error in hse_uploader: {str(e)}")
            return jsonify({'status': 'error', 'message': f"An unexpected error occurred: {str(e)}"})

    return render_template('hse_uploader.html')

def read_hse_file(file_path):
    # Updated column widths based on user input
    col_specs = [
        (0, 10),   # stock code
        (11, 13),  # issue
        (14, 22),  # required date
        (23, 31),  # required quantities
        (32, 42),  # order reference
        (43, 48),  # location
        (55, 75),  # message (adjusted to cover full width)
        (76, 91),  # last delivery note
        (92, 100)  # last delivery date
    ]

    try:
        # Read the file using read_fwf
        df = pd.read_fwf(file_path, colspecs=col_specs, header=None)
        
        # Assign column names
        df.columns = [
            'Stock Code', 'Issue', 'Required Date', 'Required Quantities',
            'Order Reference', 'Location', 'Message', 'Last Delivery Note', 'Last Delivery Date'
        ]

        # Strip whitespace from all string columns
        for col in df.select_dtypes(include=['object']):
            df[col] = df[col].str.strip()

        # Handle '00' values in the 'Issue' column
        df['Issue'] = df['Issue'].replace('00', '0')

        # Convert 'Required Date' to datetime format
        df['Required Date'] = pd.to_datetime(df['Required Date'], format='%Y%m%d', errors='coerce')

        # Convert 'Required Quantities' to string and remove leading zeros
        df['Required Quantities'] = df['Required Quantities'].astype(str).str.lstrip('0')

        # Replace empty strings with '0' for 'Required Quantities'
        df['Required Quantities'] = df['Required Quantities'].replace('', '0')

        # Filter to include all rows
        df_filtered = df.copy()

        # Exclude rows with specific stock codes
        excluded_stock_codes = ['400/U4238', '400/U4239', '331/62858']
        df_filtered = df_filtered[~df_filtered['Stock Code'].isin(excluded_stock_codes)]

        # Format the 'Month' column to include the year
        df_filtered['Month'] = df_filtered['Required Date'].dt.strftime('%B %Y')

        # Convert 'Last Delivery Date' to datetime format
        df_filtered['Last Delivery Date'] = pd.to_datetime(df_filtered['Last Delivery Date'], format='%Y%m%d', errors='coerce')

        # Convert 'Required Date' and 'Last Delivery Date' to date objects
        df_filtered['Required Date'] = df_filtered['Required Date'].dt.date
        df_filtered['Last Delivery Date'] = df_filtered['Last Delivery Date'].dt.date

        # Drop duplicate rows based on 'Stock Code', 'Issue', 'Required Date', and 'Required Quantities'
        df_filtered = df_filtered.drop_duplicates(subset=['Stock Code', 'Issue', 'Required Date', 'Required Quantities'])

        # Convert 'Required Quantities' to float
        df_filtered['Required Quantities'] = df_filtered['Required Quantities'].astype(float)

        # Add 'Unit Price' column with a default value of 0
        df_filtered['Unit Price'] = 0

        # Calculate 'Sale Price' as 'Unit Price' * 'Required Quantities'
        df_filtered['Sale Price'] = df_filtered['Unit Price'] * df_filtered['Required Quantities']

        # Replace NaN values with appropriate fill values or methods
        fill_values = {
            'Message': '',  # Replace NaN values in 'Message' column with an empty string
            'Unit Price': 0,  # Replace NaN values in 'Unit Price' column with 0
            'Sale Price': 0  # Replace NaN values in 'Sale Price' column with 0
        }
        df_filtered = df_filtered.fillna(fill_values)

        # Log the first few rows for debugging
        logging.info(f"First 5 rows of processed data:\n{df_filtered.head().to_string()}")

        return df_filtered

    except Exception as e:
        logging.error(f"Error processing file {file_path}: {str(e)}")
        raise


def get_account_id(filename):
    prefix_to_account = {
        "EMVX": "BAM002",
        "WPCX": "BAM003",
        "ATTX": "BAM004",
        "SPEX": "BAM005",
        "LPWX": "BAM007",
        "JHPX": "BAM008",
        "BHLX": "BAM009",
        "LDLX": "BAM011",
        "PWRX": "BAM018"
    }
    
    print(f"Input filename: {filename}")  # Add this line for debugging
    for prefix in prefix_to_account:
        if filename.startswith(prefix):
            account_id = prefix_to_account[prefix]
            print(f"Prefix: {prefix}, Account ID: {account_id}")  # Add this line for debugging
            return account_id

    print("Prefix not found, returning 'Unknown'")  # Add this line for debugging
    return 'Unknown'

app = create_app()

if __name__ == '__main__':
    app.run(debug=True)
    

